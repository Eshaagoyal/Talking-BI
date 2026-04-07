import os

import re

import csv

import json
import time
import logging

from io import BytesIO, StringIO

from typing import List, Optional, Tuple

from fastapi import FastAPI, HTTPException, UploadFile, File, Request

from fastapi.middleware.cors import CORSMiddleware

from pydantic import BaseModel

from dotenv import load_dotenv



import psycopg2.extras



load_dotenv()



from agents.sql_agent import (

    get_connection,

    run_sql_agent,

    get_schema,

    test_connection,

    list_datasets,

)

from agents.deep_prep import clean_and_structure

from agents.doc2chart import build_chart_data

from agents.insight_eval import evaluate_insights



app = FastAPI(title="Talking BI API")
logger = logging.getLogger("talking_bi.api")
if not logger.handlers:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s - %(message)s")


@app.middleware("http")
async def request_timing_middleware(request: Request, call_next):
    start = time.perf_counter()
    try:
        response = await call_next(request)
    except Exception:
        elapsed_ms = (time.perf_counter() - start) * 1000
        logger.exception("request_failed method=%s path=%s elapsed_ms=%.1f", request.method, request.url.path, elapsed_ms)
        raise
    elapsed_ms = (time.perf_counter() - start) * 1000
    logger.info(
        "request method=%s path=%s status=%s elapsed_ms=%.1f",
        request.method,
        request.url.path,
        response.status_code,
        elapsed_ms,
    )
    return response



_cors = (os.getenv("CORS_ORIGINS") or "*").strip()

if _cors == "*":

    app.add_middleware(

        CORSMiddleware,

        allow_origins=["*"],

        allow_credentials=False,

        allow_methods=["*"],

        allow_headers=["*"],

    )

else:

    _origins = [o.strip() for o in _cors.split(",") if o.strip()]

    app.add_middleware(

        CORSMiddleware,

        allow_origins=_origins,

        allow_credentials=False,

        allow_methods=["*"],

        allow_headers=["*"],

    )





class DashboardRequest(BaseModel):

    query: str

    kpis: List[str] = []
    num_visualizations: Optional[int] = 0
    color_schema: Optional[str] = "blue"
    preferred_chart_types: List[str] = []

    # Exact Postgres table name from GET /datasets, or "primary" for default pick

    dataset_key: Optional[str] = "primary"





class ChatRequest(BaseModel):

    message: str

    dashboard_context: dict = {}

    answer_style: Optional[str] = "balanced"  # concise | balanced | detailed

class ExplainChartRequest(BaseModel):
    title: str
    description: str
    data: list





@app.get("/")

def root():

    return {"message": "Talking BI backend is running"}





@app.get("/health")

def health():

    try:

        db = test_connection()

        groq_ok = bool(os.getenv("GROQ_API_KEY"))

        return {

            "status": "ok",

            "database": db,

            "groq_key_set": groq_ok,

            "overall": "ALL OK — ready to generate dashboards"

        }

    except Exception as e:

        raise HTTPException(status_code=500, detail=str(e))





@app.get("/datasets")

def datasets_endpoint():

    """Tables you can query — same names as in Supabase public schema."""

    try:

        return {"datasets": list_datasets()}

    except Exception as e:

        raise HTTPException(status_code=500, detail=str(e))





@app.get("/schema")

def schema_endpoint():

    try:

        return get_schema()

    except Exception as e:

        raise HTTPException(status_code=500, detail=str(e))





def _quote_ident(name: str) -> str:

    n = (name or "").strip().replace("\x00", "") or "col"

    return '"' + n.replace('"', '""') + '"'





def _parse_tabular_upload(filename: str, raw: bytes) -> Tuple[List[str], List[List[str]]]:

    fn = (filename or "data.csv").lower()

    if fn.endswith((".xlsx", ".xlsm")):

        try:

            from openpyxl import load_workbook

        except ImportError as e:

            raise ValueError("Excel support requires openpyxl (install backend dependencies).") from e

        wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)

        try:

            ws = wb[wb.sheetnames[0]]

            header: List[str] = []

            body: List[List[str]] = []

            for i, row in enumerate(ws.iter_rows(values_only=True)):

                cells = list(row)

                if i == 0:

                    header = [

                        ("" if c is None else str(c)).strip() or f"column_{j + 1}"

                        for j, c in enumerate(cells)

                    ]

                else:

                    body.append(["" if c is None else str(c) for c in cells])

        finally:

            wb.close()

        if not header:

            raise ValueError("Excel file has no rows.")

        return header, body

    if fn.endswith(".xls"):

        raise ValueError("Legacy .xls is not supported. Save as .xlsx or .csv.")

    # CSV (default)

    try:

        text = raw.decode("utf-8-sig")

    except UnicodeDecodeError:

        text = raw.decode("latin-1")

    reader = csv.reader(StringIO(text))

    try:

        header_raw = next(reader)

    except StopIteration:

        raise ValueError("File has no header row.")

    header = [(h.strip() or f"column_{i + 1}") for i, h in enumerate(header_raw)]

    body = list(reader)

    return header, body





def _write_table_from_rows(table_name: str, header: List[str], rows: List[List[str]]) -> Tuple[int, List[str]]:

    col_names = [_quote_ident(h) for h in header]

    if len(set(col_names)) != len(col_names):

        raise ValueError("Duplicate column names after normalization.")



    if len(rows) > 500_000:

        raise ValueError("File too large (max 500k data rows).")



    cols_sql = ", ".join(f"{c} TEXT" for c in col_names)

    conn = get_connection()

    cur = conn.cursor()

    try:

        cur.execute(f'DROP TABLE IF EXISTS "{table_name}" CASCADE')

        cur.execute(f'CREATE TABLE "{table_name}" ({cols_sql})')

        if rows:

            placeholders = ", ".join(["%s"] * len(col_names))

            insert_sql = f'INSERT INTO "{table_name}" ({", ".join(col_names)}) VALUES ({placeholders})'

            batch = []

            for row in rows:

                padded = list(row) + [""] * (len(header) - len(row))

                batch.append(tuple(padded[: len(header)]))

            psycopg2.extras.execute_batch(cur, insert_sql, batch, page_size=500)

        conn.commit()

        cur.execute(f'SELECT COUNT(*) FROM "{table_name}"')

        count = cur.fetchone()[0]

    except Exception as e:

        conn.rollback()

        raise

    finally:

        cur.close()

        conn.close()

    return int(count), [h.strip() for h in header]





@app.post("/upload-csv")

async def upload_csv(file: UploadFile = File(...), table_name: str = "uploaded_sales_data"):

    """

    Load CSV or Excel (.xlsx) into Postgres as a single table with the given name.

    Replaces the table if it already exists. All columns are TEXT for compatibility.

    """

    if not re.match(r"^[a-zA-Z_][a-zA-Z0-9_]*$", table_name):

        raise HTTPException(status_code=400, detail="Invalid table_name (use letters, numbers, underscore; start with letter or _).")

    raw = await file.read()

    if not raw:

        raise HTTPException(status_code=400, detail="Empty file")

    try:

        header, rows = _parse_tabular_upload(file.filename or "", raw)

    except ValueError as e:

        raise HTTPException(status_code=400, detail=str(e)) from e

    try:

        count, columns = _write_table_from_rows(table_name, header, rows)

    except Exception as e:

        raise HTTPException(status_code=500, detail=str(e)) from e



    return {"status": "ok", "table": table_name, "row_count": count, "columns": columns}





@app.post("/generate-dashboards")

def generate_dashboards(req: DashboardRequest):

    try:

        kpis = [str(k).strip().lower() for k in (req.kpis or []) if str(k).strip()]

        if not kpis:
            try:
                from agents.sql_agent import _explore_database, resolve_focus_tables
                from agents.kpi_agent import detect_kpis
                exploration = _explore_database()
                focus_tables, _ = resolve_focus_tables(exploration, req.dataset_key or "primary")
                if focus_tables and focus_tables[0] in exploration:
                    kpis = detect_kpis(req.query.strip(), exploration[focus_tables[0]])
            except Exception as e:
                print(f"Auto KPI detection error: {e}")

        num_viz = req.num_visualizations or 0



        try:

            sql_result = run_sql_agent(

                req.query.strip(),

                kpis,

                dataset_key=req.dataset_key or "primary",

            )

        except ValueError as e:

            raise HTTPException(status_code=400, detail=str(e)) from e



        if sql_result["row_count"] == 0:

            raise HTTPException(

                status_code=404,

                detail="No data returned. Try a broader query."

            )



        prep_result = clean_and_structure(

            sql_result["data"],

            req.query,

            kpis,

            num_dashboards=num_viz,

            preferred_chart_types=req.preferred_chart_types or None,

        )

        dashboard_plan = prep_result.get("dashboard_plan", {})



        if not dashboard_plan:

            raise HTTPException(

                status_code=500,

                detail="Dashboard planning failed. Try a different query."

            )



        charts = build_chart_data(sql_result["data"], dashboard_plan)

        insights = evaluate_insights(charts, kpis, req.query)

        resolved_dataset = sql_result.get("dataset_key") or (sql_result.get("focus_tables") or [None])[0]



        return {

            "status": "success",

            "query": req.query,

            "kpis": kpis,

            "dataset_key": resolved_dataset,

            "color_schema": req.color_schema,

            "sql_used": sql_result["sql"],

            "row_count": sql_result["row_count"],

            "num_dashboards": num_viz,

            "focus_tables": sql_result.get("focus_tables", []),

            "data_quality": prep_result.get("data_quality", {}),

            "dashboards": charts,

            "insights": insights,

        }



    except HTTPException:

        raise

    except Exception as e:

        raise HTTPException(status_code=500, detail=str(e))


def _extract_gemini_text(response) -> str:
    """Gemini may leave .text empty (safety, structure); read candidates/parts."""
    if response is None:
        return ""
    try:
        t = (getattr(response, "text", None) or "").strip()
        if t:
            return t
    except (ValueError, AttributeError):
        pass
    try:
        chunks = []
        for cand in getattr(response, "candidates", None) or []:
            content = getattr(cand, "content", None)
            if not content:
                continue
            for part in getattr(content, "parts", None) or []:
                tx = getattr(part, "text", None)
                if tx:
                    chunks.append(tx)
        return " ".join(chunks).strip()
    except Exception:
        return ""


@app.post("/chat")
def chat_endpoint(req: ChatRequest):
    """Chatbot — Q&A on the current dashboard or general BI questions."""
    try:
        from openai import OpenAI

        client = OpenAI(
            api_key=os.getenv("GROQ_API_KEY"),
            base_url="https://api.groq.com/openai/v1",
        )

        context = req.dashboard_context
        style = (req.answer_style or "balanced").lower().strip()
        q = str(context.get("query", ""))
        no_dashboard = q.startswith("No dashboard")

        if no_dashboard:
            guides = {
                "concise": "Answer in 1-2 clear sentences. No chart numbers required.",
                "balanced": (
                    "Answer in 3-5 sentences. For conceptual BI questions (e.g. what is data, KPI, metric), "
                    "give plain definitions and short examples. Do not invent chart numbers."
                ),
                "detailed": (
                    "Answer in one focused paragraph. Explain BI or analytics concepts when asked; use examples. "
                    "Do not require dashboard numbers."
                ),
            }
            mode_note = "MODE: No dashboard is open. Answer general BI, analytics, or conceptual questions helpfully."
        else:
            guides = {
                "concise": "Answer in 1-2 short sentences. Lead with the key number or conclusion when data supports it.",
                "balanced": "Answer in 2-4 sentences. Reference specific numbers from the dashboard when relevant.",
                "detailed": (
                    "Answer in one focused paragraph (about 4-7 sentences). Include brief context, numbers from the data "
                    "when available, and one practical implication."
                ),
            }
            mode_note = "MODE: A dashboard is open—ground answers in the context and chart snippets when the question is about this report."

        length_guide = guides.get(style, guides["balanced"])

        prompt = f"""You are a helpful BI analyst assistant.

{mode_note}

Dashboard context (may be minimal if none loaded):
- Original query: {context.get('query', 'unknown')}
- KPIs tracked: {context.get('kpis', [])}
- Insight summary: {context.get('insight_summary', 'not available')}
- Dashboard data: {json.dumps(context.get('dashboards', {}), default=str)}

User question: "{req.message}"

Instructions: {length_guide}
Always produce a helpful answer. Plain text only, no markdown."""

        response = client.chat.completions.create(
            model=os.getenv("GROQ_MODEL", "llama-3.3-70b-versatile"),
            messages=[{"role": "user", "content": prompt}]
        )
        text = response.choices[0].message.content.strip()
        if not text:
            raise HTTPException(
                status_code=502,
                detail="The model returned an empty reply. Check GROQ_API_KEY and try again.",
            )
        return {"response": text}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/explain-chart")
def explain_chart_endpoint(req: ExplainChartRequest):
    """Generates a non-technical 2-3 sentence explanation of a single chart."""
    try:
        from openai import OpenAI
        client = OpenAI(
            api_key=os.getenv("GROQ_API_KEY"),
            base_url="https://api.groq.com/openai/v1",
        )
        
        prompt = f"""You are a helpful BI assistant speaking to a non-technical user.
Please explain the following chart in 2 to 3 very simple, clear sentences. 
Highlight the biggest takeaway (like the highest value, or a trend). Do not use markdown or complex formatting.

Chart Title: {req.title}
Chart Description: {req.description}
Raw Data:
{json.dumps(req.data, default=str)[:3000]}
"""
        response = client.chat.completions.create(
            model=os.getenv("GROQ_MODEL", "llama-3.3-70b-versatile"),
            messages=[{"role": "user", "content": prompt}]
        )
        text = response.choices[0].message.content.strip()
        if not text:
            text = "Sorry, I couldn't generate an explanation for this chart right now."
        return {"explanation": text}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

